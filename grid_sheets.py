"""
GRID v2 — Sheets & Excel Integration
Google Sheets (read/write/create) + local Excel files (.xlsx) + data analytics.
Shares OAuth2 credentials with grid_google.
"""

import json
import os
import re
from datetime import datetime
from typing import Optional

from grid_google import _load_config, _save_config, _get_credentials

SHEET_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
]

PLOTS_DIR = "plots"


# ── helpers ──────────────────────────────────────────────────────

def _ensure_plots_dir():
    os.makedirs(PLOTS_DIR, exist_ok=True)


def _fmt_table(headers: list, rows: list, max_rows: int = 50) -> str:
    if not rows:
        return "(no data)"
    display = rows[:max_rows]
    col_widths = [len(h) for h in headers]
    for r in display:
        for i, v in enumerate(r):
            col_widths[i] = max(col_widths[i], len(str(v)))
    sep = "  ".join("-" * w for w in col_widths)
    line = "  ".join(h.ljust(w) for h, w in zip(headers, col_widths))
    out = [line, sep]
    for r in display:
        out.append("  ".join(str(v).ljust(w) for v, w in zip(r, col_widths)))
    if len(rows) > max_rows:
        out.append(f"... and {len(rows) - max_rows} more rows")
    return "\n".join(out)


# ── SheetsClient (Google Sheets API) ────────────────────────────

class SheetsClient:
    def __init__(self):
        self.service = None

    def _ensure(self):
        if self.service:
            return
        creds = _get_credentials()
        if not creds:
            raise RuntimeError("Google not authenticated. Run /gcal setup first.")
        from googleapiclient.discovery import build
        self.service = build("sheets", "v4", credentials=creds)
        self.drive = build("drive", "v3", credentials=creds)

    def list_sheets(self, max_results: int = 20) -> str:
        self._ensure()
        q = "mimeType='application/vnd.google-apps.spreadsheet'"
        result = self.drive.files().list(q=q, pageSize=max_results,
                                         fields="files(id,name,modifiedTime)").execute()
        files = result.get("files", [])
        if not files:
            return "No Google Sheets found."
        lines = [f"Google Sheets ({len(files)}):"]
        for f in files:
            t = f.get("modifiedTime", "?")[:10]
            lines.append(f"  {f['name']}  [{f['id']}]  (modified {t})")
        return "\n".join(lines)

    def read(self, file_id: str, range_: str = "Sheet1") -> str:
        self._ensure()
        result = self.service.spreadsheets().values().get(
            spreadsheetId=file_id, range=range_).execute()
        values = result.get("values", [])
        if not values:
            return "Sheet is empty."
        headers = values[0]
        rows = values[1:]
        return _fmt_table(headers, rows) + f"\n\n{len(rows)} rows x {len(headers)} cols"

    def create(self, title: str) -> str:
        self._ensure()
        from googleapiclient.discovery import build
        body = {"properties": {"title": title}}
        sheet = self.service.spreadsheets().create(body=body).execute()
        sid = sheet["spreadsheetId"]
        url = sheet.get("spreadsheetUrl", f"https://docs.google.com/spreadsheets/d/{sid}")
        return f"Created: {title}\n  ID: {sid}\n  URL: {url}"

    def write(self, file_id: str, range_: str, values: list) -> str:
        self._ensure()
        body = {"values": values}
        result = self.service.spreadsheets().values().update(
            spreadsheetId=file_id, range=range_,
            valueInputOption="USER_ENTERED", body=body).execute()
        return f"Wrote {result.get('updatedCells', 0)} cells to {range_}."

    def append(self, file_id: str, range_: str, values: list) -> str:
        self._ensure()
        body = {"values": values}
        result = self.service.spreadsheets().values().append(
            spreadsheetId=file_id, range=range_,
            valueInputOption="USER_ENTERED", body=body).execute()
        return f"Appended {result.get('updates', {}).get('updatedCells', 0)} cells."


# ── Local Excel (.xlsx) ─────────────────────────────────────────

def _load_xlsx(path: str):
    from openpyxl import load_workbook
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path)


def excel_open(path: str, sheet_name: Optional[str] = None, max_rows: int = 50) -> str:
    wb = _load_xlsx(path)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return f"Sheet '{sheet_name}' is empty."
    headers = [str(c) if c is not None else "" for c in rows[0]]
    data_rows = [[str(c) if c is not None else "" for c in r] for r in rows[1:]]
    out = [f"File: {os.path.basename(path)}  Sheet: {sheet_name}"]
    out.append(_fmt_table(headers, data_rows, max_rows))
    out.append(f"{len(data_rows)} rows x {len(headers)} cols")
    return "\n".join(out)


def excel_create(path: str, sheet_name: str = "Sheet1",
                 headers: Optional[list] = None) -> str:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if headers:
        ws.append(headers)
    wb.save(path)
    return f"Created: {path}  (sheet: {sheet_name})"


def excel_edit_cell(path: str, cell: str, value: str, sheet_name: Optional[str] = None) -> str:
    wb = _load_xlsx(path)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    ws[cell] = value
    wb.save(path)
    return f"Set {cell} = '{value}' in {os.path.basename(path)}"


def excel_add_row(path: str, values: list, sheet_name: Optional[str] = None) -> str:
    wb = _load_xlsx(path)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    ws.append(values)
    wb.save(path)
    return f"Row added to {os.path.basename(path)}"


# ── Data Analytics ──────────────────────────────────────────────

def _data_to_df(source: str, sheet: Optional[str] = None):
    """Load data from Google Sheet ID or local .xlsx path into a pandas DataFrame."""
    import pandas as pd
    if source.startswith("http"):
        source = source.split("/d/")[1].split("/")[0] if "/d/" in source else source
    if "/" in source or os.path.exists(source):
        return pd.read_excel(source, sheet_name=sheet or 0)
    # Treat as Google Sheet ID
    creds = _get_credentials()
    if not creds:
        raise RuntimeError("Google not authenticated.")
    sid = source
    url = f"https://docs.google.com/spreadsheets/d/{sid}/export?format=csv"
    return pd.read_csv(url)


def data_stats(input_str: str) -> str:
    parts = input_str.strip().split(maxsplit=2)
    source = parts[0] if parts else ""
    sheet = parts[1] if len(parts) > 1 else None
    if not source:
        return "Usage: stats <sheet_id_or_path> [sheet_name]"
    try:
        df = _data_to_df(source, sheet)
        if df.empty:
            return "No data loaded."
        desc = df.describe(include="all").to_string()
        info = f"{len(df)} rows x {len(df.columns)} cols\nColumns: {', '.join(df.columns)}"
        return f"{info}\n\nSummary:\n{desc}"
    except Exception as e:
        return f"Error: {e}"


def data_analyze(input_str: str) -> str:
    parts = input_str.strip().split(maxsplit=2)
    source = parts[0] if parts else ""
    if not source:
        return "Usage: analyze <sheet_id_or_path> [columns...]"
    col_filter = parts[1] if len(parts) > 1 else ""
    try:
        import pandas as pd
        df = _data_to_df(source)
        if df.empty:
            return "No data loaded."
        cols = [c.strip() for c in col_filter.split(",")] if col_filter else list(df.columns)
        out = []
        for c in cols:
            if c not in df.columns:
                out.append(f"  Column '{c}' not found.")
                continue
            s = df[c]
            out.append(f"--- {c} ---")
            out.append(f"  Type: {s.dtype}")
            out.append(f"  Count: {s.count()}  Missing: {s.isna().sum()}")
            if s.dtype in ("object", "string"):
                out.append(f"  Unique: {s.nunique()}")
                top = s.value_counts().head(5)
                for v, cnt in top.items():
                    out.append(f"    '{v}': {cnt}")
            else:
                out.append(f"  Min: {s.min()}  Max: {s.max()}  Mean: {s.mean():.2f}")
                out.append(f"  Median: {s.median()}  Std: {s.std():.2f}")
                q = s.quantile([0.25, 0.5, 0.75])
                out.append(f"  Q1: {q[0.25]:.2f}  Median: {q[0.5]:.2f}  Q3: {q[0.75]:.2f}")
        return "\n".join(out)
    except Exception as e:
        return f"Error: {e}"


def data_plot(input_str: str) -> str:
    parts = input_str.strip().split("|")
    source = parts[0].strip() if parts else ""
    if not source:
        return "Usage: plot <sheet_id_or_path> | <x_col> | <y_col> [| title]"
    x_col = parts[1].strip() if len(parts) > 1 else ""
    y_col = parts[2].strip() if len(parts) > 2 else ""
    title = parts[3].strip() if len(parts) > 3 else f"{y_col} vs {x_col}"
    if not source or not x_col or not y_col:
        return "Usage: plot <sheet_id_or_path> | <x_col> | <y_col> [| title]"
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import pandas as pd
        df = _data_to_df(source)
        if x_col not in df.columns:
            return f"Column '{x_col}' not found."
        if y_col not in df.columns:
            return f"Column '{y_col}' not found."
        _ensure_plots_dir()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"{PLOTS_DIR}/plot_{ts}.png"
        df.plot(x=x_col, y=y_col, kind="line", title=title, rot=45)
        plt.tight_layout()
        plt.savefig(fname)
        plt.close()
        return f"Plot saved: {fname}"
    except Exception as e:
        return f"Error: {e}"


# ═══════════════════════════════════════════════════════════════════
# Tool function
# ═══════════════════════════════════════════════════════════════════

def _gsheet_help() -> str:
    return """Sheets & Excel commands (prefix with tool name or use /gsheet):

  Google Sheets (requires OAuth2 setup via /gcal setup):
    list                              — List your Google Sheets
    read <id> [range]                 — Read data from a sheet (range defaults to Sheet1)
    create <title>                    — Create a new Google Sheet
    write <id> | <range> | <val1,...> — Write row(s) to a sheet
    append <id> | <range> | <val1,...> — Append row(s)

  Local Excel (.xlsx):
    open <path> [sheet]               — Open and display a local .xlsx file
    new <path> [sheet_name] [hdr,...]  — Create a new .xlsx file
    edit <path> | <cell> | <value>    — Edit a cell in a local .xlsx
    addrow <path> | <val1,...>        — Add a row to a local .xlsx

  Data Analytics (works with sheet ID or local path):
    stats <id_or_path> [sheet]        — Summary statistics of the data
    analyze <id_or_path> [columns]    — Deep column-by-column analysis
    plot <id_or_path> | <x> | <y> [| title] — Plot two columns (saves as PNG)

  help                                — Show this help

  Identifiers: Google Sheet ID = long string from the sheet URL, or full URL.
  Pipe | separates arguments within a command."""


def google_sheets(input_str: str) -> str:
    parts = input_str.strip().split(maxsplit=1)
    cmd = parts[0].lower() if parts else "help"
    args = parts[1] if len(parts) > 1 else ""

    if cmd == "help":
        return _gsheet_help()

    # ── Local Excel commands (no Google auth needed) ──────────

    if cmd == "open":
        pa = args.split(maxsplit=1)
        path = pa[0] if pa else ""
        sheet = pa[1] if len(pa) > 1 else None
        if not path:
            return "Usage: open <path> [sheet_name]"
        return excel_open(path, sheet)

    if cmd == "new":
        pa = args.split(maxsplit=2)
        path = pa[0] if pa else ""
        sheet = pa[1] if len(pa) > 1 else "Sheet1"
        hdrs = [h.strip() for h in pa[2].split(",")] if len(pa) > 2 and pa[2] else None
        if not path:
            return "Usage: new <path> [sheet_name] [header1,header2,...]"
        return excel_create(path, sheet, hdrs)

    if cmd == "edit":
        if "|" not in args:
            return "Usage: edit <path> | <cell> | <value>"
        pa = [x.strip() for x in args.split("|", 2)]
        path = pa[0]
        cell = pa[1] if len(pa) > 1 else ""
        value = pa[2] if len(pa) > 2 else ""
        if not path or not cell:
            return "Usage: edit <path> | <cell> | <value>"
        return excel_edit_cell(path, cell, value)

    if cmd == "addrow":
        if "|" not in args:
            return "Usage: addrow <path> | <val1,val2,...>"
        pa = [x.strip() for x in args.split("|", 1)]
        path = pa[0]
        vals = [v.strip() for v in pa[1].split(",")] if len(pa) > 1 else []
        if not path or not vals:
            return "Usage: addrow <path> | <val1,val2,...>"
        return excel_add_row(path, vals)

    # ── Data analytics (works with both) ──────────────────────

    if cmd == "stats":
        return data_stats(args)

    if cmd == "analyze":
        return data_analyze(args)

    if cmd == "plot":
        return data_plot(args)

    # ── Google Sheets commands (require auth) ─────────────────

    try:
        sc = SheetsClient()
    except Exception as e:
        return f"Google auth error: {e}\nRun /gcal setup first."

    if cmd == "list":
        return sc.list_sheets()

    if cmd == "read":
        ra = args.split(maxsplit=1)
        fid = ra[0] if ra else ""
        range_ = ra[1] if len(ra) > 1 else "Sheet1"
        if not fid:
            return "Usage: read <sheet_id> [range]"
        return sc.read(fid, range_)

    if cmd == "create":
        if not args.strip():
            return "Usage: create <title>"
        return sc.create(args.strip())

    if cmd == "write":
        if "|" not in args:
            return "Usage: write <id> | <range> | <val1,...>"
        pa = [x.strip() for x in args.split("|", 2)]
        fid = pa[0]
        range_ = pa[1] if len(pa) > 1 else "Sheet1"
        raw = pa[2] if len(pa) > 2 else ""
        if not fid or not raw:
            return "Usage: write <id> | <range> | <val1,...>"
        rows = [[v.strip() for v in row.split(",")] for row in raw.split(";")]
        return sc.write(fid, range_, rows)

    if cmd == "append":
        if "|" not in args:
            return "Usage: append <id> | <range> | <val1,...>"
        pa = [x.strip() for x in args.split("|", 2)]
        fid = pa[0]
        range_ = pa[1] if len(pa) > 1 else "Sheet1"
        raw = pa[2] if len(pa) > 2 else ""
        if not fid or not raw:
            return "Usage: append <id> | <range> | <val1,...>"
        rows = [[v.strip() for v in row.split(",")] for row in raw.split(";")]
        return sc.append(fid, range_, rows)

    return f"Unknown command: {cmd}\n{_gsheet_help()}"
